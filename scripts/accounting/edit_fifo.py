# %% Prepare
import os
from string import ascii_uppercase as alphabeth

import dotenv
import pandas as pd
from openpyxl import load_workbook, Workbook

dotenv.load_dotenv()

# %% Constants

DATABASE_URL = os.getenv("FIFO_DB")
with open("scripts/accounting/dea_transactions", newline="", encoding="utf-8") as file:
    QUERY = file.read()
OUTPUT_FILE = "data/statements/Olek/fifo/FIFO_05.2026.xlsx"
# %%
ROW_START = 15
COL_BDATE = 3
COL_BNUMBER = 4
COL_BQUANTITY = 5
COL_BWORTH = 6
COL_BPRICE = 7
COL_SDATE = 10
COL_SNUMBER = 11
COL_SQUANTITY = 12
COL_SWORTH = 13
# %%


def prepare():
    print("Fetching transactions...")
    df = pd.read_sql_query(QUERY, DATABASE_URL)

    df["timestamp"] = pd.to_datetime(df["timestamp"])
    if hasattr(df["timestamp"].dt, "tz") and df["timestamp"].dt.tz is not None:
        df["timestamp"] = df["timestamp"].dt.tz_localize(None)

    print(f"Found {len(df)} transactions")
    return df


def edit_sheet(writer: Workbook, asset_df: pd.DataFrame, asset: str):
    sheet_name = str(asset).upper().replace("/", "_").replace(".", "_").replace(" ", "").replace("°", "o")
    try:
        ws = writer[f"FIFO_{sheet_name}"]
    except:
        print(f"Skipping '{sheet_name}' ({len(asset_df)} txns)")
        return

    for _, tx in asset_df.iterrows():
        idx = tx["position"] + ROW_START - 1
        tx["currency"] = tx["currency"].replace(" ", "").replace("°", "o")
        if tx["type"].lower() == "buy":
            date = COL_BDATE
            nr = COL_BNUMBER
            qty = COL_BQUANTITY
            val = (
                f"='FIFO_{tx['currency']}'!{alphabeth[COL_SWORTH - 1]}{int(tx['linked_position'] or 0)}"
                if tx["reversed"]
                else None
            )
            ws.cell(row=idx, column=COL_BWORTH, value=val or tx["total"])
            ws.cell(row=idx, column=COL_BPRICE, value=f"={alphabeth[COL_BWORTH - 1]}{idx}/{alphabeth[qty - 1]}{idx}")
        else:
            date = COL_SDATE
            nr = COL_SNUMBER
            qty = COL_SQUANTITY

        ws.cell(row=idx, column=date, value=tx["timestamp"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(
            row=idx,
            column=nr,
            value=tx["external_id"] or f"Wymiana Spot {tx['asset']}/{tx['currency']} {tx['exchange']}",
        )
        ws.cell(row=idx, column=qty, value=abs(tx["quantity"]))

    print(f"Created '{sheet_name}' ({len(asset_df)} txns)")


def main():
    df = prepare()
    writer = load_workbook(OUTPUT_FILE)

    for asset in sorted(df["asset"].unique()):
        edit_sheet(writer, df[df["asset"] == asset].sort_values(["type", "timestamp"]).reset_index(drop=True), asset)

    writer.save("fifo.xlsx")
    writer.close()

    print(f"✅ Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
