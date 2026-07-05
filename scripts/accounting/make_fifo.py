# %% Prepare
import os
from datetime import datetime
from string import ascii_uppercase as alphabeth

import dotenv
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

dotenv.load_dotenv()

# %% Constants

DATABASE_URL = os.getenv("FIFO_DB")
with open("scripts/accounting/dea_transactions.sql", newline="", encoding="utf-8") as file:
    QUERY = file.read()
OUTPUT_FILE = "fifo.xlsx"

ROW_LEGEND = 1
ROW_QTY = 2
ROW_COST = 3
ROW_AVG = 4

OFFSET = 12
COL_LEGEND = 1 + OFFSET
COL_OPENING = 2 + OFFSET
COL_BUY = 3 + OFFSET
COL_SELL = 4 + OFFSET
COL_DIFF = 5 + OFFSET
COL_CLOSE = 6 + OFFSET

L_OPENING = alphabeth[COL_OPENING - 1]
L_BUY = alphabeth[COL_BUY - 1]
L_SELL = alphabeth[COL_SELL - 1]
L_DIFF = alphabeth[COL_DIFF - 1]

headers = [
    "Time",
    "Type",
    "ID",
    "Qty",
    "Total_Value",
    "Fee",
    "Price",
    "Running_Balance",
    "Cumulative_Cost",
    "Order ID",
]
COL_TYPE = alphabeth[headers.index("Type")]
COL_ID = alphabeth[headers.index("ID")]
COL_QTY = alphabeth[headers.index("Qty")]
COL_VALUE = alphabeth[headers.index("Total_Value")]
COL_FEE = alphabeth[headers.index("Fee")]
COL_PRICE = alphabeth[headers.index("Price")]
COL_BALANCE = alphabeth[headers.index("Running_Balance")]
COL_CUM = alphabeth[headers.index("Cumulative_Cost")]
COL_NOTE = alphabeth[headers.index("Order ID")]
ROW_START = 2

# %%


def make_legend(ws: Worksheet, last_row: int, asset_idx: int):
    ws.cell(row=ROW_QTY, column=COL_LEGEND, value="Qty:")
    ws.cell(row=ROW_COST, column=COL_LEGEND, value="Total Cost:")
    ws.cell(row=ROW_AVG, column=COL_LEGEND, value="Avg Cost:")

    ws.cell(row=ROW_LEGEND, column=COL_OPENING, value="OPENING:")
    ws.cell(row=ROW_QTY, column=COL_OPENING, value=f"='_SETTINGS_'!B{asset_idx + 2}")
    ws.cell(row=ROW_COST, column=COL_OPENING, value=f"='_SETTINGS_'!C{asset_idx + 2}")

    ws.cell(row=ROW_LEGEND, column=COL_BUY, value="BUY")
    ws.cell(
        row=ROW_QTY,
        column=COL_BUY,
        value=f'=SUMIF({COL_TYPE}{ROW_START}:{COL_TYPE}{last_row},"BUY",{COL_VALUE}{ROW_START}:{COL_VALUE}{last_row})',
    )
    ws.cell(
        row=ROW_COST,
        column=COL_BUY,
        value=f'=SUMIFS({COL_PRICE}{ROW_START}:{COL_PRICE}{last_row},{COL_TYPE}{ROW_START}:{COL_TYPE}{last_row},"BUY")',
    )

    ws.cell(row=ROW_LEGEND, column=COL_SELL, value="SELL")
    ws.cell(
        row=ROW_QTY,
        column=COL_SELL,
        value=f'=SUMIF({COL_TYPE}{ROW_START}:{COL_TYPE}{last_row},"SELL",{COL_VALUE}{ROW_START}:{COL_VALUE}{last_row})',
    )
    ws.cell(
        row=ROW_COST,
        column=COL_SELL,
        value=f'=SUMIFS({COL_PRICE}{ROW_START}:{COL_PRICE}{last_row},{COL_TYPE}{ROW_START}:{COL_TYPE}{last_row},"SELL")',
    )

    ws.cell(row=ROW_LEGEND, column=COL_DIFF, value="Diff")
    ws.cell(row=ROW_QTY, column=COL_DIFF, value=f"=SUM({L_BUY}{ROW_QTY};{L_SELL}{ROW_QTY})")
    ws.cell(row=ROW_COST, column=COL_DIFF, value=f"=SUM({L_BUY}{ROW_COST};{L_SELL}{ROW_COST})")

    ws.cell(row=ROW_LEGEND, column=COL_CLOSE, value="CLOSING:")
    ws.cell(row=ROW_QTY, column=COL_CLOSE, value=f"={COL_BALANCE}{last_row}")
    ws.cell(row=ROW_COST, column=COL_CLOSE, value=f"={COL_CUM}{last_row}")
    ws.cell(row=ROW_AVG, column=COL_CLOSE, value=f"={COL_CUM}{last_row}/{COL_BALANCE}{last_row}")
    ws.add_table(
        Table(
            displayName=f"Summaries_{asset_idx}",
            ref="M1:R4",
            tableStyleInfo=TableStyleInfo("TableStyleMedium9", showFirstColumn=True),
        )
    )

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)


def create_dashboard(writer: pd.ExcelWriter, assets):
    # ===== DASHBOARD =====
    ws_dash = writer.book.create_sheet("DASHBOARD")
    ws_dash.cell(row=1, column=1, value="Portfolio Overview")

    ws_dash.cell(row=3, column=1, value="Asset")
    ws_dash.cell(row=3, column=2, value="Closing_Qty")
    ws_dash.cell(row=3, column=3, value="Closing_Cost($)")

    for i, asset in enumerate(assets, start=4):
        sheet = str(asset).upper().replace("/", "_").replace(".", "_").replace(" ", "")[:31]
        ws_dash.cell(row=i, column=1, value=asset)
        ws_dash.cell(row=i, column=2, value=f"='{sheet}'!F2")
        ws_dash.cell(row=i, column=3, value=f"='{sheet}'!F3")

    ws_dash.cell(row=len(assets) + 6, column=1, value="TOTAL:")
    ws_dash.cell(row=len(assets) + 6, column=3, value=f"=SUM(C4:C{len(assets) + 3})")

    ws_dash.cell(row=len(assets) + 8, column=1, value="Generated:")
    ws_dash.cell(row=len(assets) + 8, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


def prepare():
    print("Fetching transactions...")
    df = pd.read_sql_query(QUERY, DATABASE_URL)

    df["timestamp"] = pd.to_datetime(df["timestamp"])
    if hasattr(df["timestamp"].dt, "tz") and df["timestamp"].dt.tz is not None:
        df["timestamp"] = df["timestamp"].dt.tz_localize(None)

    print(f"Found {len(df)} transactions")

    writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

    # ===== SHEET 1: RAW_DATA =====
    df.to_excel(writer, sheet_name="RAW_DATA", index=False)
    print("Written RAW_DATA")
    return df, writer


def create_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, asset: str, asset_idx: int):
    sheet_name = str(asset).upper().replace("/", "_").replace(".", "_").replace(" ", "")[:31]
    ws = writer.book.create_sheet(sheet_name)

    # ===== OPENING BALANCE + SUMMARIES =====
    asset_df = df[df["asset"] == asset].sort_values("timestamp").reset_index(drop=True)
    make_legend(ws, len(asset_df), asset_idx)

    for idx, (_, tx) in enumerate(asset_df.iterrows(), start=ROW_START):
        ws.cell(row=idx, column=headers.index("Time") + 1, value=str(tx["timestamp"]))
        ws.cell(row=idx, column=headers.index("Type") + 1, value=tx["type"].upper())
        ws.cell(row=idx, column=headers.index("Qty") + 1, value=tx["quantity"])
        ws.cell(row=idx, column=headers.index("Total_Value") + 1, value=tx["total"])
        ws.cell(row=idx, column=headers.index("Fee") + 1, value=tx.get("fee", 0) or 0)
        ws.cell(row=idx, column=headers.index("Price") + 1, value=f"={COL_VALUE}{idx}/ABS({COL_QTY}{idx})")
        ws.cell(
            row=idx,
            column=headers.index("ID") + 1,
            value=str(tx.get("external_id")) or f"Spot {tx.get('asset')}/{tx.get('currency')}",
        )

        if idx == ROW_START:
            ws.cell(row=idx, column=headers.index("Running_Balance") + 1, value=f"={L_OPENING}2+{COL_QTY}{idx}")
            prev_cum = f"{L_OPENING}3"  # Opening cost from SETTINGS cell
            prev_bal = f"{L_OPENING}2"
        else:
            ws.cell(
                row=idx,
                column=headers.index("Running_Balance") + 1,
                value=f"={COL_BALANCE}{idx - 1}+{COL_QTY}{idx}",
            )
            prev_cum = f"{COL_CUM}{idx - 1}"
            prev_bal = f"{COL_BALANCE}{idx - 1}"

        if tx["type"].lower() == "buy":
            cum_formula = f"={prev_cum}+(ABS({COL_QTY}{idx})*{COL_VALUE}{idx})"
        else:
            avg_cost_formula = f"IF({prev_bal}=0,0,{prev_cum}/{prev_bal})"
            cum_formula = f"={prev_cum}-(ABS({COL_QTY}{idx})*{avg_cost_formula})"

        ws.cell(row=idx, column=headers.index("Cumulative_Cost") + 1, value=cum_formula)

    ws.add_table(
        Table(
            displayName=f"Transactions_{asset_idx}",
            ref=f"A1:{alphabeth[len(headers) - 1]}{len(asset_df)}",
            tableStyleInfo=TableStyleInfo("TableStyleMedium2", showRowStripes=True),
        )
    )

    print(f"Created '{sheet_name}' ({len(asset_df)} txns)")


def main():
    df, writer = prepare()

    assets = sorted(df["asset"].unique())

    # ===== SHEET 2: _SETTINGS_ =====
    settings_data = [["Asset", "Opening_Qty", "Opening_Total_Cost"]]
    for asset in assets:
        settings_data.append([asset, "", ""])
    settings_df = pd.DataFrame(settings_data[1:], columns=["Asset", "Opening_Qty", "Opening_Total_Cost"])
    settings_df.to_excel(writer, sheet_name="_SETTINGS_", index=False)

    # ===== SHEETS 3+: Per Currency =====
    for asset_idx, asset in enumerate(assets):
        create_sheet(writer, df, asset, asset_idx)

    create_dashboard(writer, assets)

    writer.close()

    print(f"✅ Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
